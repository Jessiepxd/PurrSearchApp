from docx2python import docx2python  # pip install docx2python python-docx
import openpyxl  # pip install openpyxl
import fitz  # pip install PyMuPDF
from vsdx import VisioFile # pip install vsdx
from docx import Document
import xlrd

import gzip
import os
import re

import sys
print(sys.version)

def xls_search(file_path, text):
    """Search inside .xls files and return matches with context."""
    matches = []
    try:
        # Open the .xls file
        workbook = xlrd.open_workbook(file_path)
        text_lower = text.lower()

        # Iterate through all sheets
        for sheet in workbook.sheets():
            for row_num in range(sheet.nrows):
                for col_num in range(sheet.ncols):
                    cell_value = str(sheet.cell_value(row_num, col_num)).lower()
                    if text_lower in cell_value:
                        # Capture some context (surrounding rows)
                        matches.append(f"Sheet: '{sheet.name}' | Row: {row_num + 1} | Cell: {col_num + 1} | Value: {cell_value}")
                        if len(matches) >= 10:
                            matches.append(f"[... more results in this sheet]")
                            return matches

    except Exception as e:
        print(f"Error processing {file_path}: {e}")

    return matches



def is_log_file(file_path):
    """Determine if a file is a log file based on naming patterns."""
    file_name = os.path.basename(file_path)

    # Handle gzipped files
    if file_name.endswith('.gz'):
        file_name = file_name[:-3]  # Remove .gz extension

    # Check for exact filenames without extensions
    if file_name.startswith(('messages', 'syslog', 'dmesg')):
        return True

    # Check if ".log" appears anywhere in the filename
    if '.log' in file_name.lower():
        return True

    # Check if extension is a number (e.g., .1, .2, etc.)
    extension = os.path.splitext(file_name)[1]  # Get extension including the dot
    last_part = extension.rpartition('.')[2]    # Get part after the last dot
    if last_part.isdecimal():
        return True

    # Check for specific patterns that wouldn't be caught by previous checks
    if file_name.startswith('dmesg.') or file_name.startswith('apt.txt.'):
        return True

    return False

def test_is_log_file():
    """Test the is_log_file function with various filenames and expected results."""
    test_cases = {
        # Original test cases
        'messages': True,
        'dmesg': True,
        'syslog': True,
        "messages.": True,
        'messages.gz': True,
        'dmesg.gz': True,
        'syslog.gz': True,
        'messages.33': True,
        'myfile.33': True,
        'myfile.33.gz': True,
        'dpkg.log.20191201.1575181021.gz': True,
        'apt.txt.20221212.0': True,
        'dmesg.0': True,
        'dmesg.1.gz': True,
        'dpkg.log': True,
        'dpkg.log.1': True,
        'dpkg.log.20191201.1575181021.gz': True,
        'dpkg.log.20220902.1662135421': True,
        'mcadapter.log.Fri': True,
        'turbine.log-2020-02-28-10-1.gz': True,
        'wsrunner_RenderXBrowserAPI.log-2019-03-15-13-1.gz': True,
        'messages.99.gz': True,
        'syslog.20220811.1660267021.gz': True,
        'jessie.blog': False
    }

    temp_dir = "/tmp/"

    print("Testing is_log_file function:")
    print("-" * 70)
    print("Filename                            | Result | Expected | Match?")
    print("-" * 70)

    all_passed = True

    for file_name, expected in test_cases.items():
        file_path = os.path.join(temp_dir, file_name)
        result = is_log_file(file_path)
        match = result == expected
        if not match:
            all_passed = False
        print(f"{file_name:35} | {result:6} | {expected:8} | {match}")

    print("-" * 70)
    if all_passed:
        print("All tests passed!")
    else:
        print("Some tests failed. Check your implementation.")


def log_search(file_path, text, context_chars=32, case_sensitive=False):
    """Search in log files, handling gzipped files if necessary."""
    matches = []
    search_text = text if case_sensitive else text.lower()

    try:
        # Check file magic bytes to determine if it's gzipped
        is_gzipped = False
        with open(file_path, 'rb') as test_file:
            magic_bytes = test_file.read(2)
            if magic_bytes == b'\x1f\x8b':  # GZip magic number
                is_gzipped = True

        # Choose the appropriate file opener based on whether the file is gzipped
        opener = gzip.open if is_gzipped else open
        # open_mode = 'rt' if is_gzipped else 'r'

        # Use the selected opener with a single code path for both file types
        with opener(file_path, 'rt', encoding='utf-8', errors='ignore') as file:
            for line_num, line in enumerate(file, start=1):
                line_text = line if case_sensitive else line.lower()
                if search_text in line_text:
                    # Highlight the matched text within the line
                    highlighted_line = line.replace(text, f"[MATCH: {text}]")
                    matches.append(f"Line {line_num}: {highlighted_line.strip()}")

        return matches
    except Exception as e:
        return [f"Error processing {file_path}: {str(e)}"]


def binary_search(file_path, text, context_chars=32):
    matches = []
    text_lower = text.lower()  # Convert search text to lowercase for case-insensitive search

    # Open and read the file as text
    with open(file_path, 'rt', encoding='utf8', errors='ignore') as file:
        data = file.read().lower()  # Convert the entire content to lowercase for case-insensitive matching

    pos = data.find(text_lower)  # Find the first occurrence
    while pos >= 0:
        # Extract a portion of text around the match for context
        start = max(pos - context_chars, 0)
        end = min(pos + len(text_lower) + context_chars, len(data))
        context = data[start:end]  # Get the context string

        # Highlight the matched text within the context
        highlighted_match = context.replace(text_lower, f"[MATCH: {text}]")

        # Append the context with the matched text
        matches.append(f"CONTEXT: '{highlighted_match}'")

        # Find the next occurrence
        pos = data.find(text_lower, pos + len(text_lower))

    # return "\n".join(matches)
    return matches


# .doc, .dot
def mbcs_search(file_path, text):
    """Search for raw text encoded as UTF-16."""
    matches = []
    utf16_text = text.lower().encode('utf-16')[2:]

    with open(file_path, 'rb') as file:
        data = file.read().lower()

    pos = data.find(utf16_text)
    while pos >= 0:
        matches.append({"position": pos})
        pos = data.find(utf16_text, pos + len(utf16_text))

    return matches

def combined_search(file_path, text):
    """Perform both mbcs_search and binary_search on the file."""
    matches_mbcs = mbcs_search(file_path, text)
    matches_binary = binary_search(file_path, text)

    # Combine the results and remove duplicates
    combined_matches = list({match['position']: match for match in matches_mbcs + matches_binary}.values())

    # return sorted(combined_matches)
    return sorted(combined_matches, key=lambda x: x['position'])


# .docm
def extract_text_from_docx_element(element):
    """
    Recursively extracts text from a docx2python element,
    handling nested lists and concatenating text content.
    """
    if isinstance(element, list):
        return '\n'.join([extract_text_from_docx_element(item) for item in element])
    elif isinstance(element, str):
        return element
    else:
        return ''

def docm_python_search(file_path, text):
    try:
        result = docx2python(file_path)
        # Extracting text properly from the docx2python result object, handling nested structures
        text_content = extract_text_from_docx_element(result.body)
        text_content = text_content.lower()
        matches = []
        start = 0
        while start != -1:
            start = text_content.find(text.lower(), start)
            if start != -1:
                matches.append({"position": start})
                start += len(text)
        return matches
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []


# .docx
""" return paragraph number """
def docx_python_search(file_path, text):
    try:
        doc = Document(file_path)
        matches = []
        search_text_lower = text.lower()

        for i, paragraph in enumerate(doc.paragraphs):
            if search_text_lower in paragraph.text.lower():
                matches.append(f"Paragraph: {i + 1} | Text: {paragraph.text}")

        return matches
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []

""" return page number (failed)"""

def xlsx_search(file_path, text) -> list[str]:
    """Search XLSX file and return a list with descriptions of each match, including cell content."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        matches = []
        text_lower = text.lower()
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and text_lower in str(cell.value).lower():
                        matches.append(f"Sheet: '{sheet_name}' | Cell: {cell.coordinate} | Content: {cell.value}")
        return matches
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []


def pdf_search(file_path, text, context_chars=32):
    try:
        doc = fitz.open(file_path)  # Open the PDF document
        matches = []
        text_lower = text.lower()  # Convert search text to lowercase

        # Iterate through each page of the document
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)  # Load the page
            page_text = page.get_text("text").lower()  # Get all text content from the page in lowercase

            # Search for all occurrences of the term in the page
            start_pos = page_text.find(text_lower)
            while start_pos != -1:
                # Extract context around the match
                start_context = max(0, start_pos - context_chars)
                end_context = min(len(page_text), start_pos + len(text_lower) + context_chars)
                context = page_text[start_context:end_context]

                # Append the match along with the page number
                matches.append(f"Page {page_num + 1}:\n{context.strip()}\n")

                # Find the next occurrence on the page
                start_pos = page_text.find(text_lower, start_pos + len(text_lower))

        doc.close()  # Close the document
        # print("\n".join(matches))
        return matches  # Return all matches found
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []



def vsdx_search(file_path, search_text):
    matches = set()  # Using a set to store unique page names

    with VisioFile(file_path) as visio:
        # Function to recursively search through shapes and their sub-shapes
        def search_shapes(shapes, page_name):
            for shape in shapes:
                shape_text = shape.text
                if shape_text and search_text.lower() in shape_text.lower():
                    matches.add(page_name)  # Add the page name to the set
                    return  # Stop searching this page as we've found a match

                # Recursively search sub-shapes
                if shape.sub_shapes():
                    search_shapes(shape.sub_shapes(), page_name)

        # Iterate through all pages in the Visio file
        for page in visio.pages:
            search_shapes(page.shapes, page.name)

    # Convert the set to a list of dictionaries
    found_items = [page for page in matches]

    return found_items


if __name__ == "__main__":
    # Execute test functions if module is run.
    test_is_log_file()