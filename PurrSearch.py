import os
import sys
import collections
import time
import images
import locale
import threading
import wx.adv
import wx
import queue
import subprocess
from wx.lib.throbber import Throbber
import datetime
from search_functions import binary_search, combined_search, docm_python_search, docx_python_search, xlsx_search, pdf_search, vsdx_search, xls_search, is_log_file, log_search
# import wx.richtext
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# wxPython GUI Framework ----------------------------------------------
try:
    import wx          # wxPython - Requires 4.0+ or newer.
    import images
except ImportError as err:
    print(str(err))
    print("wxPython required: http://www.wxpython.org")
    print("python -m pip install wxpython")
    sys.exit(1)
# Third-Party Imports -------------------------------------------------
try:
    import openpyxl    # OpenPyXL for reading/writing .XLSX files.
except ImportError as err:
    raise ImportError('\n'.join((
        str(err), "Requires OpenPyXL: https://openpyxl.readthedocs.io/en/stable/",
        "\tpip install openpyxl")))
if not openpyxl.DEFUSEDXML:
    # OpenPyXL uses defusedxml to protect against XML attacks.
    raise ImportError('\n'.join((
        "Requires defusedxml: https://pypi.org/project/defusedxml/",
        "\tpip install defusedxml")))
if not openpyxl.LXML:
    # LXML uses efficient C libraries libxml2 and libxslt for speed.
    raise ImportError('\n'.join(("Requires lxml: https://lxml.de/",
                                "\tpip install lxml")))
try:
    import fitz     # fitz (PyMuPDF)
except ImportError as err:
    raise ImportError('\n'.join((
        str(err), "PyMuPDF required: https://pypi.org/project/PyMuPDF/",
        "\tpython -m pip install pymupdf")))

try:
    from vsdx import VisioFile     # vsdx library
except ImportError as err:
    raise ImportError('\n'.join((
        str(err), "vsdx required: https://pypi.org/project/vsdx/",
        "\tpython -m pip install vsdx")))

# Application defines --------------------------------------------------
APPNAME = "PurrSearch"
VENDORNAME = "Evertz"
COMPANYNAME = "Evertz Microsystems Ltd."
PRODUCTNAME = "For Purr-fect Searches"
COPYRIGHT = "2024 Evertz Microsystems Ltd."

# Version information for Windows "VersionInfo" resource.
VERSION = "3.9.9"            # First functional build.
VERSIONPARTS = [int(part) for part in VERSION.split('.')]
VERSIONPARTS += [0] * max(0, 4 - len(VERSIONPARTS))
VERSIONINFO = {
    'fileMajor': VERSIONPARTS[0], 'fileMinor': VERSIONPARTS[1],
    'fileMicro': VERSIONPARTS[2], 'fileBuild': VERSIONPARTS[3],
    'prodMajor': VERSIONPARTS[0], 'prodMinor': VERSIONPARTS[1],
    'prodMicro': VERSIONPARTS[2], 'prodBuild': VERSIONPARTS[3],
    'companyName': COMPANYNAME, 'fileDescription': PRODUCTNAME,
    'fileVersion': VERSION, 'internalName': APPNAME,
    'legalCopyright': COPYRIGHT, 'originalFilename': APPNAME + '.exe',
    'productName': PRODUCTNAME, 'productVersion': VERSION,
    'comments': '', 'legalTrademarks': '', }


class SearchThread(threading.Thread):
    """Thread to perform searches on all files in a particular folder."""
    def __init__(self, parent, path, text, result_queue):
        super().__init__()
        self.parent = parent
        self.path = path
        self.text = text
        self.result_queue = result_queue

        self.initial_message = None
        self.progress = 0
        self.statusText = "Status: Not Running"
        self.error = None

    def run(self):
        if self.parent.end_event.is_set(): # means the search is set to stop
            return
        try:
            self.performSearch(self.path, self.text)
        except Exception as err:
            self.error = f"Search failed with error: {err}"
            self.result_queue.put(('error', self.error))

    def performSearch(self, path, text):
        start_time = time.time()  # Start timing the search
        processed_files = 0
        files_with_matches = 0  # Counter for files that have matches
        total_files, self.initial_message = self.parent.calculateTotalFiles(path)

        self.result_queue.put(('status', self.initial_message))
        self.result_queue.put(('progress', self.progress, "0%"))

        # Implement the search across files, skipping temporary and unrelated file formats
        skipped_formats = ['.jpg', '.db', '.png', '.wbk', '.jpeg', '.pptx', '.shs', '.lnk', '.tmp', '.bmp', '.msg', '.vsd']
        for root, dirs, files in os.walk(path):
            for file in files:
                if self.parent.end_event.is_set():
                    self.result_queue.put(('status', "\n\n\u1360 Search stopped by user.\u1360\n"))
                    self.result_queue.put(('progress', 100, "Stopped"))
                    return

                if file.startswith('~$'):
                    continue

                file_path = os.path.join(root, file)
                extension = os.path.splitext(file_path)[1].lower()
                self.result_queue.put(('status_bar', f"Searching in file: {file}"))

                if extension in skipped_formats:
                    continue

                try:
                    matches = []
                    # Check if it's a log file (including gzipped logs)
                    if is_log_file(file_path):
                        matches = log_search(file_path, text)
                    elif extension in ['.txt', '.rtf', '.csv', '.mib', '.bat', '.sh', '.c', '.cpp', '.h', '.cs', 'html', '.htm', '.css', '.php', '.js', '.xml', '.ini', '.cfg', '.json', '.java', '.tex', '.rst', '.md', '.ps', '.nfo', '.info', '.py', '.yaml', '.toml']:
                        matches = binary_search(file_path, text)
                    elif extension == '.xls':
                        matches = xls_search(file_path, text)
                    elif extension in ['.doc', '.dot']:
                        matches = combined_search(file_path, text)
                    elif extension == '.docx':
                        matches = docx_python_search(file_path, text)
                    elif extension == '.docm':
                        matches = docm_python_search(file_path, text)
                    elif extension == '.xlsx':
                        matches = xlsx_search(file_path, text)
                    elif extension == '.pdf':
                        matches = pdf_search(file_path, text)
                    elif extension == '.vsdx':
                        matches = vsdx_search(file_path, text)


                    if matches:
                        result_text = "\n".join(matches)
                        self.result_queue.put(('result', file, len(matches), result_text, file_path))
                        files_with_matches += 1
                except Exception as e:
                    self.error = f"\n\nError processing {file_path}: {e}\n\n"
                    self.result_queue.put(('error', self.error))

                # Update the progress bar and percentage
                processed_files += 1
                self.progress = int((processed_files / total_files) * 100)
                self.result_queue.put(('progress', self.progress, f"{self.progress}%"))

        # Calculate the total search time, reset the search button
        elapsed_time = time.time() - start_time
        self.initial_message = f"\n\nSearch finished in {elapsed_time:.3f} seconds. {files_with_matches} files found with matches."
        self.result_queue.put(('status', self.initial_message))
        self.result_queue.put(('complete',))
        self.parent.end_event.set()

    def stop(self, block=True):
        """Signal the thread to end."""
        self.parent.end_event.set()
# end class SearchThread(threading.Thread)


class DetailsDialog(wx.Dialog):
    """Simple dialog to show search results."""
    def __init__(self, parent, details: str):
        super().__init__(parent, title="Details", size=(600, 500),
                         style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER)
        # Dialog layout
        sizer = wx.BoxSizer(wx.VERTICAL)
        # Format the details to be displayed line by line
        formatted_details = details.replace(" | ", "\n")
        # TextCtrl for showing the details
        detailsCtrl = wx.TextCtrl(self, value=formatted_details,
                                  style=wx.TE_MULTILINE | wx.TE_READONLY |
                                  wx.HSCROLL | wx.TE_RICH)
        sizer.Add(detailsCtrl, proportion=1, flag=wx.EXPAND | wx.ALL, border=10)

        buttons = self.CreateStdDialogButtonSizer(wx.OK)
        sizer.Add(buttons, flag=wx.EXPAND | wx.ALL, border=0)
        self.SetSizer(sizer)
# end class DetailsDialog(wx.Dialog)


class ResultsList(wx.ListCtrl):
    """Virtual List Control to show search results."""
    # Columns as (title, alignment, width)
    COLUMNS = (('File', wx.LIST_FORMAT_LEFT, 300),
               ('Matches', wx.LIST_FORMAT_LEFT, 80),
               ('Details', wx.LIST_FORMAT_LEFT, 380),
               ('Path', wx.LIST_FORMAT_LEFT, 400))

    def __init__(self, parent, style=0):
        super().__init__(parent, style=style | wx.LC_VIRTUAL | wx.LC_REPORT |
                         wx.LC_SINGLE_SEL)
        self.items: list[tuple[str, str, str, str]] = []
        for num, (title, alignment, width) in enumerate(self.COLUMNS):
            self.InsertColumn(num, title, format=alignment, width=width)

    def OnGetItemText(self, item, column):
        """Return the text to be displayed for the given item (row) and column"""
        return self.items[item][column]
# end class ResultsList(wx.ListCtrl)


class MainPanel(wx.Panel):
    def __init__(self, parent, wx_config: wx.ConfigBase):
        super().__init__(parent)
        self.search_thread = None
        self.end_event = threading.Event()
        self.result_queue = queue.Queue()

        self.initial_layout_done = False  # Flag to control when to adjust column width

        # Add an accelerator table to capture the F1 key
        self.setupAcceleratorTable()

        sizer = wx.BoxSizer(wx.VERTICAL)
        self.wx_config = wx_config  # wx.Config() object.
        self.current_path = None

        # Timer for updating UI
        self.timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.onTimer, self.timer)
        self.timer.Start(100)

        # Layout UI ===================================================
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.wx_config = wx_config  # wx.Config() object.
        self.current_path = None

        # Directory Selection -----------------------------------------
        hbox1 = wx.BoxSizer(wx.HORIZONTAL)

        label = wx.StaticText(self, label="Select a folder:")
        self.current_path = self.wx_config.Read('/lastpath', defaultVal='')
        self.dirCtrl = wx.TextCtrl(self, value=self.current_path, style=wx.TE_READONLY)
        self.browseBtn = wx.Button(self, label="Browse")
        self.browseBtn.SetBitmap(images.browse_16.GetBitmap())

        hbox1.Add(label, flag=wx.ALIGN_CENTER_VERTICAL)
        hbox1.AddSpacer(10)
        hbox1.Add(self.dirCtrl, proportion=1)
        hbox1.AddSpacer(10)
        hbox1.Add(self.browseBtn)

        # Text Search -------------------------------------------------
        hbox2 = wx.BoxSizer(wx.HORIZONTAL)
        label = wx.StaticText(self, label="Enter text to search for:")
        self.search_history = self.loadSearchHistory()

        self.searchCtrl = wx.ComboBox(self, choices=self.search_history + ["*** Clear Search & Path History ***"], style=wx.CB_DROPDOWN  | wx.TE_PROCESS_ENTER)
        self.searchCtrl.Bind(wx.EVT_COMBOBOX, self.onComboBoxSelect)
        self.searchCtrl.Bind(wx.EVT_TEXT_ENTER, self.onStartSearch)

        # Auto-select the last search term if the history is not empty
        if self.search_history:
            self.searchCtrl.SetValue(self.search_history[0])  # Select the most recent search term

        self.startSearchBtn = wx.Button(self, label="Search")
        self.stopSearchBtn = wx.Button(self, label="Stop")
        self.stopSearchBtn.Disable()

        hbox2.Add(label, flag=wx.ALIGN_CENTER_VERTICAL)
        hbox2.AddSpacer(10)
        hbox2.Add(self.searchCtrl, proportion=1)
        hbox2.AddSpacer(10)
        hbox2.Add(self.startSearchBtn)
        hbox2.AddSpacer(10)
        hbox2.Add(self.stopSearchBtn)

        # Progress Bar ------------------------------------------
        hbox3 = wx.BoxSizer(wx.HORIZONTAL)

        self.progressBar = wx.Gauge(self, range=100, style=wx.GA_HORIZONTAL)
        self.progressLabel = wx.StaticText(self, label="0%", size=(50, -1))

        imgs = [images.catalog.get(f"frame_{i}_delay-0.03s").GetBitmap() for i in range(90)]
        self.throbber = Throbber(self, -1, imgs, frameDelay=0.03)
        self.throbber.Stop()

        # Load the search done image (same as browse_16)
        self.search_done_bitmap = wx.StaticBitmap(self, bitmap=images.complete_16.GetBitmap())
        self.search_done_bitmap.Hide()  # Hide it initially

        hbox3.Add(self.progressBar, proportion=1, flag=wx.ALIGN_CENTER_VERTICAL)
        hbox3.AddSpacer(10)
        hbox3.Add(self.throbber, flag=wx.ALIGN_CENTER_VERTICAL)
        hbox3.Add(self.search_done_bitmap, flag=wx.ALIGN_CENTER_VERTICAL)
        hbox3.AddSpacer(10)
        hbox3.Add(self.progressLabel, flag=wx.ALIGN_CENTER_VERTICAL)


        # Results Display
        splitter = wx.SplitterWindow(self, style=wx.SP_LIVE_UPDATE)
        self.resultsCtrl = wx.TextCtrl(splitter, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)

        self.resultsList = ResultsList(splitter, style=wx.LC_HRULES | wx.LC_VRULES)

        self.resultsList.Bind(wx.EVT_LIST_ITEM_ACTIVATED, self.onItemActivated)
        self.resultsList.Bind(wx.EVT_CONTEXT_MENU, self.onRightClick)
        #######evt_context_menu

        self.Bind(wx.EVT_SIZE, self.onResize)

        splitter.SplitHorizontally(self.resultsCtrl, self.resultsList)
        splitter.SetMinimumPaneSize(50)

        # Information Button -------------------------------------------
        hbox4 = wx.BoxSizer(wx.HORIZONTAL)

        self.statusText = "Status: Ready to search"
        self.statusLabel = wx.StaticText(self, label=self.statusText)  # Initial status
        hbox4.Add(self.statusLabel, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT, border=10)

        self.exportBtn = wx.Button(self, label="Export")

        infoBtn = wx.Button(self, label="About")
        infoBtn.SetBitmap(images.info_16.GetBitmap())

        hbox4.AddStretchSpacer(1)
        hbox4.Add(self.exportBtn)
        hbox4.Add(infoBtn)


        # Arrange All Sections
        sizer.Add(hbox1, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(hbox2, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(hbox3, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(splitter, proportion=1, flag=wx.EXPAND | wx.ALL, border=10)
        sizer.Add(hbox4, flag=wx.EXPAND | wx.RIGHT | wx.BOTTOM, border=10)
        self.SetSizer(sizer)

        # Bind All Events
        self.browseBtn.Bind(wx.EVT_BUTTON, self.onBrowse)
        self.startSearchBtn.Bind(wx.EVT_BUTTON, self.onStartSearch)
        self.stopSearchBtn.Bind(wx.EVT_BUTTON, self.onStopSearch)
        infoBtn.Bind(wx.EVT_BUTTON, self.onInfo)
        self.exportBtn.Bind(wx.EVT_BUTTON, self.onExportResults)

    def setupAcceleratorTable(self):
            """Setup accelerator table to capture F1 key and open About dialog."""
            accel_tbl = wx.AcceleratorTable([
                (wx.ACCEL_NORMAL, wx.WXK_F1, wx.ID_HELP)
            ])
            self.SetAcceleratorTable(accel_tbl)

            # Bind the F1 key to the onInfo method
            self.Bind(wx.EVT_MENU, self.onInfo, id=wx.ID_HELP)

    def onItemActivated(self, event):
        index = event.GetIndex()
        details = self.resultsList.items[index][2]
        self.showDetailsDialog(details)

    def showDetailsDialog(self, details):
        # Create the dialog window
        dlg = DetailsDialog(self, details)
        dlg.ShowModal()  # Show the dialog as modal
        dlg.Destroy()    # Destroy the dialog after it's closed`

    def onComboBoxSelect(self, event):
        """Handle the selection of an item in the combobox."""
        selection = self.searchCtrl.GetValue()
        if selection == "*** Clear Search & Path History ***":
            self.clearSearchHistory()
            self.searchCtrl.SetValue('')

    def clearSearchHistory(self):
        """Clears the search history from the configuration and the combobox."""
        self.search_history.clear()  # Clear the history list
        for i in range(10):  # Assuming 10 is the max number of history items
            self.wx_config.Write(f'/searchHistory/item{i}', '')  # Clear each entry in the config

        self.wx_config.Write('/lastpath', '')
        self.current_path = ''
        self.dirCtrl.SetValue('')

        self.wx_config.Flush()  # Write changes to disk
        self.searchCtrl.Clear()  # Clear the combobox entries
        self.searchCtrl.AppendItems(["*** Clear Search & Path History ***"])  # Add only the clear option
        self.searchCtrl.SetValue('')  # Reset the displayed value

    # Button Functions ================================================
    def onBrowse(self, event: wx.CommandEvent):
        """User clicked Browse button, show directory picker dialog."""
        with wx.DirDialog(self, "Choose a folder", style=wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as dirDialog:
            if dirDialog.ShowModal() == wx.ID_CANCEL:
                return
            self.current_path = dirDialog.GetPath()
            self.dirCtrl.SetValue(self.current_path)
            self.wx_config.Write('/lastpath', self.current_path)
            # HKEY_CURRENT_USER/Software/Evertz/PurrSearch/lastpath

    def onInfo(self, event):
        """Showing app details."""
        info = wx.adv.AboutDialogInfo()
        info.SetName(APPNAME)
        info.SetVersion(VERSION)
        info.SetDescription(
            f"{PRODUCTNAME}\n\n"
            f"Company: {COMPANYNAME}\n"
            f"\u00A9 {COPYRIGHT} All rights reserved.\n\n"
            f"Supported file types:\n"
            f".txt, .rtf, .xls, .doc, .dot, .docx, .docm, .xlsx, .pdf, .vsdx, .csv\n"
            f".mib, .bat, .sh, .c, .cpp, .h, .cs, .html, .htm, .css, .php, .py, .info\n"
            f".js, .xml, .ini, .cfg, .json, .java, .tex, .rst, .md, .ps, .nfo, .yaml, .toml\n"
        )
        info.SetWebSite("https://evertz.com/")

        icon_bitmap = images.icon_64.GetBitmap()
        icon = wx.Icon()
        icon.CopyFromBitmap(icon_bitmap)
        info.SetIcon(icon)

        wx.adv.AboutBox(info)

    def calculateTotalFiles(self, path):
        total_files = 0
        file_types = collections.Counter()
        skipped_formats = ['.jpg', '.db', '.png', '.wbk', '.jpeg', '.pptx', '.shs', '.lnk', '.tmp', '.bmp', '.msg', '.vsd']
        log_count = 0

        for root, dirs, files in os.walk(path):
            for file in files:
                if file.startswith('~$'):  # Skip temporary files
                    continue
                file_path = os.path.join(root, file)
                extension = os.path.splitext(file)[1].lower()

                # Check if it's a log file first
                if is_log_file(file_path):
                    total_files += 1
                    log_count += 1
                    file_types['log'] += 1
                elif extension == '.gz' and is_log_file(file_path[:-3]):
                    total_files += 1
                    log_count += 1
                    file_types['log.gz'] += 1
                elif extension not in skipped_formats:
                    total_files += 1
                    file_types[extension] += 1

        self.initial_message = f"Searching in {total_files} files.\nThis may take a while depending on the number of files.\n\n"
        self.initial_message += "FILES COUNT BY TYPE:"

        for ext, count in file_types.items():
            self.initial_message += f"\n{ext if ext else 'No extension'}: {count}"

        # Add log file count information
        if log_count > 0:
            self.initial_message += f"\n\nTotal log files (including compressed): {log_count}"

        return total_files, self.initial_message

    def loadSearchHistory(self):
        """Load search history from the configuration."""
        history = []
        for i in range(10):  # Load up to 10 items from history
            entry = self.wx_config.Read(f'/searchHistory/item{i}', defaultVal='')
            if entry:
                history.append(entry)
        return history

    # regedit > Computer\HKEY_CURRENT_USER\SOFTWARE\Evertz\PurrSearch\searchHistory
    def saveSearchHistory(self, term):
        """Save the search term to the configuration using enumerate."""
        if term in self.search_history:
            self.search_history.remove(term)
        self.search_history.insert(0, term)
        self.search_history = self.search_history[:10]  # Keep only the last 10 entries

        for index, entry in enumerate(self.search_history):
            self.wx_config.Write(f'/searchHistory/item{index}', entry)
        self.wx_config.Flush()  # Ensure the config is written to disk
        self.refreshSearchHistory()

    def refreshSearchHistory(self):
        """Refresh the combobox with the search history and 'Clear Search & Path History' option."""
        # self.searchCtrl.Clear()  # Clear the current items in the combobox
        self.searchCtrl.AppendItems(self.search_history + ["*** Clear Search & Path History ***"])  # Add history and clear option
        # self.searchCtrl.SetValue('')  # Optionally reset the displayed value

    def onStartSearch(self, event):
        # Check if a search is already in progress
        if self.search_thread is not None and self.search_thread.is_alive():
            return

        search_term = self.searchCtrl.GetValue()

        # Check if the directory path or search term is empty
        if not self.dirCtrl.GetValue() or not search_term:
            self.error = "Please select a folder and enter a search term."
            wx.MessageBox(self.error, "Error", wx.OK | wx.ICON_ERROR)
            return

        # Check if there are results from a previous search
        if self.resultsList.GetItemCount() > 0:
            confirm_dialog = wx.MessageDialog(
                self,
                "Old search results will be cleared. Continue?",
                "Confirm",
                wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION
            )
            if confirm_dialog.ShowModal() != wx.ID_YES:
                confirm_dialog.Destroy()
                return
            confirm_dialog.Destroy()

        self.saveSearchHistory(search_term)
        self.searchCtrl.Set(self.search_history)
        self.searchCtrl.SetValue(search_term)

        self.end_event.clear()
        self.resultsList.DeleteAllItems()  # Clear previous search results
        self.resultsList.items = []
        self.resultsList.SetItemCount(0)
        self.resultsCtrl.SetValue("")
        search_term = self.searchCtrl.GetValue()

        # Reset the throbber and hide the 'search done' image
        self.search_done_bitmap.Hide()  # Hide the search done image
        self.throbber.Show()  # Show the throbber
        self.throbber.Start()  # Start the throbber animation

        self.searchCtrl.Disable()
        self.startSearchBtn.Disable()

        # self.statusText = "Status: now searching "
        self.statusLabel.SetLabel(self.statusText)
        self.Layout()

        # Start the new search thread
        self.search_thread = SearchThread(self, self.current_path, search_term, self.result_queue)
        self.search_thread.start()

        # Start the throbber animation for the new search
        self.throbber.Start()

    def onStopSearch(self, event):
        if not self.end_event.is_set():
            self.end_event.set()
            self.statusText = "Status: Stopping..."
            # self.progressLabel.SetLabel(self.statusText)
            self.statusLabel.SetLabel(self.statusText)

            if self.search_thread.is_alive():
                self.search_thread.join()
                self.search_thread.stop()
                self.statusText = "Status: Search Stopped"
                self.statusLabel.SetLabel(self.statusText)
            self.end_event=threading.Event()
        self.searchCtrl.Enable()
        self.startSearchBtn.Enable()
        self.exportBtn.Enable()
        self.throbber.Stop()
        # self.search_thread = None

    def onSearchComplete(self):
        """Called when the search thread completes its task."""
        self.end_event.set()
        self.statusText = "Status: Search completed."
        self.statusLabel.SetLabel(self.statusText)
        self.Layout()
        self.refreshSearchHistory()
        self.stopSearchBtn.Disable()
        self.startSearchBtn.Enable()
        self.throbber.Stop()
        self.search_thread = None
        self.searchCtrl.Enable()
        self.startSearchBtn.Enable()
        self.exportBtn.Enable()

    def onTimer(self, event: wx.TimerEvent):
        """Update UI with search thread results, if thread is running."""
        if self.search_thread:
            if self.search_thread.is_alive():
                self.stopSearchBtn.Enable()
                self.startSearchBtn.Disable()
                self.browseBtn.Disable()
                self.exportBtn.Disable()
            else:
                self.stopSearchBtn.Disable()
                # if self.end_event.is_set():
                self.startSearchBtn.Enable()
                self.browseBtn.Enable()
                self.exportBtn.Enable()

            while not self.result_queue.empty():
                item = self.result_queue.get()
                if item[0] == 'status':
                    self.resultsCtrl.AppendText(item[1])
                elif item[0] == 'status_bar':
                    self.statusLabel.SetLabel(item[1])
                elif item[0] == 'progress':
                    self.progressBar.SetValue(item[1])
                    self.progressLabel.SetLabel(item[2])

                    # If progress is 100%, show the new image
                    if item[1] == 100:
                        # Stop the throbber and replace it with the search done image
                        self.throbber.Stop()
                        self.throbber.Hide()
                        # Show the "search done" bitmap
                        self.search_done_bitmap.Show()
                        self.Layout()  # Re-layout the panel to reflect changes

                        self.progressLabel.SetLabel("Search Complete!")  # Update label to show it's done

                elif item[0] == 'result':
                    self.resultsList.items.append((item[1], str(item[2]), item[3], item[4]))
                    self.resultsList.SetItemCount(len(self.resultsList.items))
                elif item[0] == 'error':
                    self.resultsCtrl.AppendText(item[1])
                elif item[0] == 'complete':
                    self.onSearchComplete()
        else:
            # If there's no search_thread, it means no search is in progress
            self.stopSearchBtn.Disable()
            self.startSearchBtn.Enable()
            self.browseBtn.Enable()

            if self.resultsList.GetItemCount() > 0:
                self.exportBtn.Enable()
            else:
                self.exportBtn.Disable()

    def onClose(self, event: wx.CloseEvent):
        """Program is closing, ask user for confirmation only if a search was performed."""
        # Check if there are any search results or if a search is in progress
        if self.search_thread is None or not self.search_thread.is_alive():
            # No search has been performed or no results are present, close without confirmation
            self.timer.Stop()
            event.Skip()  # Allow the window to close
            self.Destroy()  # Close the window directly
        else:
            # A search has been performed or results exist, ask for confirmation
            reply = wx.MessageBox("Are you sure you want to exit?",
                                "Confirm Exit", wx.YES_NO | wx.ICON_QUESTION)
            if reply == wx.NO:
                event.Veto()  # Prevent the window from closing
                return
            else:
                # Ensure that the search thread is stopped before closing
                self.end_event.set()
                if self.search_thread is not None:
                    self.search_thread.join()  # Wait for the thread to finish
            self.timer.Stop()
            event.Skip()  # Allow the window to close

    def exportResultsToExcel(self, filepath):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Search Results"

        # Get the number of columns in resultsList
        num_columns = self.resultsList.GetColumnCount()

        # Write header row
        for col in range(num_columns):
            header = self.resultsList.GetColumn(col).GetText()
            sheet.cell(row=1, column=col+1, value=header)

        # Write data rows
        for row in range(self.resultsList.GetItemCount()):
            for col in range(num_columns):
                cell_value = self.resultsList.GetItem(row, col).GetText()
                sheet.cell(row=row+2, column=col+1, value=cell_value)

        # Adjust column widths
        for col in range(1, num_columns+1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

        # Save the workbook
        workbook.save(filepath)

        reply = wx.MessageBox(f"Results successfully exported to {filepath}. Open file?", "Export Successful", wx.YES_NO | wx.ICON_INFORMATION)
        # Open the file immediately if they clicked yes.
        if reply in (wx.YES, wx.ID_YES):
            wx.SafeYield()
            if sys.platform.startswith('win'):
                os.startfile(filepath)

    def onExportResults(self, event):
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        file_name_prefix = "search_results"  # Replace with actual type if available
        default_filename = f"{file_name_prefix}_{current_date}.xlsx"

        with wx.FileDialog(self, "Save Excel file", wildcard="Excel files (*.xlsx)|*.xlsx",
                        style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT, defaultFile=default_filename) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return  # the user changed their mind

            # Save the current contents in the file
            pathname = fileDialog.GetPath()
            try:
                self.exportResultsToExcel(pathname)
            except IOError:
                wx.LogError("Cannot save current data in file '%s'." % pathname)

    def onRightClick(self, event: wx.ContextMenuEvent):
        menu = wx.Menu()
        open_file_item = wx.MenuItem(menu, wx.ID_ANY, "Open File")
        open_path_item = wx.MenuItem(menu, wx.ID_ANY, "Open Path")
        menu.Append(open_file_item)
        menu.Append(open_path_item)
        self.Bind(wx.EVT_MENU, self.onOpenFile, open_file_item)
        self.Bind(wx.EVT_MENU, self.onOpenFilePath, open_path_item)
        self.PopupMenu(menu)
        menu.Destroy()

    def onOpenFile(self, event):
        selected_item_index = self.resultsList.GetFirstSelected()
        if selected_item_index != wx.NOT_FOUND:
            file_path = self.resultsList.GetItemText(selected_item_index, 3)
            if os.path.exists(file_path):
                wx.LaunchDefaultApplication(file_path)

    def onOpenFilePath(self, event):
        selected_item_index = self.resultsList.GetFirstSelected()
        if selected_item_index != wx.NOT_FOUND:
            file_path = self.resultsList.GetItemText(selected_item_index, 3)
            directory = os.path.dirname(file_path)
            if os.path.exists(directory):
                if sys.platform.startswith('win'):
                    os.startfile(directory)
                elif sys.platform.startswith('darwin'):
                    subprocess.call(['open', '--', directory])
                else:
                    subprocess.call(['xdg-open', directory])

    def onResize(self, event):
        self.Layout()
        if self.initial_layout_done:
            self.adjust_last_column_width()
        else:
            self.initial_layout_done = True
        event.Skip()

    def adjust_last_column_width(self):
        total_width = self.resultsList.GetClientSize().GetWidth()
        other_columns_width = sum([self.resultsList.GetColumnWidth(col) for col in range(self.resultsList.GetColumnCount() - 1)])
        last_column_width = total_width - other_columns_width - 20  # 20 is a buffer for the scrollbar
        self.resultsList.SetColumnWidth(self.resultsList.GetColumnCount() - 1, last_column_width)

    def onKeyPress(self, event):
        """Handles key press events and checks for the F1 key."""
        key_code = event.GetKeyCode()
        if key_code == wx.WXK_F1:  # Check if F1 is pressed
            self.onInfo(None)  # Call the About dialog
        else:
            event.Skip()  # Continue processing other key events


class SearchGUI(wx.Frame):
    def __init__(self):
        super().__init__(parent = None, title = "PurrSearch --- For Purr-fect Searches", size = (1240, 600))
        # Load application config from wx.Config (HKCU\Software\VENDOR\APPNAME).
        self.wx_config = wx.Config()

        # Instantiate the MainPanel as a child of this Frame.
        self.panel = MainPanel(self, self.wx_config)
        self.panel.SetBackgroundColour(wx.Colour(231,223,221))

        # Setup application icon bundle
        self.icons = wx.IconBundle()
        self.icons.AddIcon(images.icon_16.GetIcon())
        self.icons.AddIcon(images.icon_32.GetIcon())
        self.icons.AddIcon(images.icon_64.GetIcon())
        self.icons.AddIcon(images.icon_128.GetIcon())
        self.icons.AddIcon(images.icon_256.GetIcon())
        self.SetIcons(self.icons)

        self.Center()


        # Create a splash screen using the embedded image
        splash_bitmap = images.catalog['splash'].GetBitmap()
        splash = wx.adv.SplashScreen(
            splash_bitmap,
            wx.adv.SPLASH_CENTRE_ON_SCREEN | wx.adv.SPLASH_TIMEOUT,
            3000,
            None
        )

        splash.Show()

        # Setup a timer to close the splash screen and show the main window
        self.splash_timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.onCloseSplash, self.splash_timer)
        self.splash_timer.Start(3000, wx.TIMER_ONE_SHOT)

    def onCloseSplash(self, event):
        self.splash_timer.Stop()  # Stop the timer
        self.Show()  # Show the main window

        self.Bind(wx.EVT_CLOSE, self.panel.onClose) ###

if __name__ == "__main__":
    import wx.lib.mixins.inspection as wit
    app = wit.InspectableApp(redirect=False)
    locale.setlocale(locale.LC_ALL, 'C')
    app.SetAppName(APPNAME)
    app.SetVendorName(VENDORNAME)
    frame = SearchGUI()
    app.MainLoop()


## add files in generate.bat
## run generate.bat
## remove the new images.py to the right folder, replace the old one
## python --version
## python -m pip install pyinstaller
## python -m PyInstaller --version
## python -m site --user-site
## Add to PATH: "Edit the system environment variables"
## pyinstaller --onefile --windowed --icon=PurrSearch.ico --add-data "PurrSearch.ico;." PurrSearch.py

## pyinstaller --onefile --windowed --icon=PurrIcon.ico --add-data "PurrIcon.ico;." PurrSearch.py
## "C:\Users\jessiec\AppData\Local\Programs\Python\Python311\Scripts\pyinstaller.exe" --onefile --windowed --icon=PurrIcon.ico --add-data "PurrIcon.ico;." PurrSearch.py
## "C:\Users\jessiec\AppData\Local\Programs\Python\Python311\Scripts\pyinstaller.exe" --onefile --windowed --icon=PurrIcon.ico --add-data "PurrIcon.ico;." --hidden-import=wx._xml PurrSearch.py

## "C:\Users\jessiec\AppData\Local\Programs\Python\Python313\Scripts\pyinstaller.exe" --onefile --windowed --icon=PurrIcon.ico --add-data "PurrIcon.ico;." PurrSearch.py


## Copy the spec file to replace current one.
## pyinstaller PurrSearch.spec
## "C:\Users\jessiec\AppData\Local\Programs\Python\Python311\Scripts\pyinstaller.exe" PurrSearch.spec